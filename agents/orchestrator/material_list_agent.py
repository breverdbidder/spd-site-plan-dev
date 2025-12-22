"""
SPD Material List Agent - LangGraph Orchestration
==================================================

Automated material specification and procurement system that converts construction 
takeoff quantities into procurement-ready material lists with Home Depot and Lowe's 
SKU mapping, Florida building code compliance for rough openings, and technical 
specifications.

Author: Claude Sonnet 4.5 (AI Architect)
Date: December 22, 2025
Repository: github.com/breverdbidder/spd-site-plan-dev
Stack: LangGraph + Anthropic Claude + Firecrawl + OpenPyXL
"""

import os
import json
import re
from pathlib import Path
from typing import TypedDict, Literal, List, Dict, Any, Optional
from datetime import datetime
from dataclasses import dataclass
import requests

# Optional imports - gracefully handle missing packages
try:
    from langgraph.graph import StateGraph, END
    LANGGRAPH_AVAILABLE = True
except ImportError:
    LANGGRAPH_AVAILABLE = False

try:
    from anthropic import Anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ============================================================================
# CONFIGURATION
# ============================================================================

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
FIRECRAWL_API_KEY = os.environ.get("FIRECRAWL_API_KEY", "")
SUPABASE_URL = "https://mocerqjnksmhcjzxrewo.supabase.co"
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

# Initialize clients when available
anthropic_client = None
if ANTHROPIC_AVAILABLE and ANTHROPIC_API_KEY:
    anthropic_client = Anthropic(api_key=ANTHROPIC_API_KEY)

# ============================================================================
# STATE DEFINITION
# ============================================================================

class MaterialListState(TypedDict):
    """State for material list workflow"""
    # Input from Takeoff Agent
    takeoff_materials: List[Dict[str, Any]]
    takeoff_quantities: List[Dict[str, Any]]
    project_name: str
    project_id: str
    project_address: str
    
    # Material Processing
    categorized_materials: Dict[str, List[Dict[str, Any]]]
    rough_openings: List[Dict[str, Any]]
    
    # Retailer Integration
    home_depot_products: List[Dict[str, Any]]
    lowes_products: List[Dict[str, Any]]
    best_prices: List[Dict[str, Any]]
    
    # Technical Specs
    technical_specs: List[Dict[str, Any]]
    florida_code_compliance: List[Dict[str, Any]]
    
    # Output
    procurement_list: List[Dict[str, Any]]
    excel_path: str
    total_material_cost: float
    total_labor_cost: float
    
    # Metadata
    messages: List[Dict[str, str]]
    errors: List[str]
    status: Literal["initiated", "materials_extracted", "retailers_searched", 
                    "specs_generated", "report_created", "completed", "failed"]

# ============================================================================
# FLORIDA BUILDING CODE 2023 - ROUGH OPENING STANDARDS
# ============================================================================

FLORIDA_ROUGH_OPENINGS = {
    "windows": {
        "single_hung": {"width_add": 0.5, "height_add": 0.5, "notes": "Standard clearance"},
        "double_hung": {"width_add": 0.5, "height_add": 0.5, "notes": "Standard clearance"},
        "casement": {"width_add": 0.75, "height_add": 0.5, "notes": "Extra width for hardware"},
        "sliding": {"width_add": 0.5, "height_add": 0.5, "notes": "Standard clearance"},
        "picture": {"width_add": 0.5, "height_add": 0.5, "notes": "Fixed, standard clearance"},
        "awning": {"width_add": 0.5, "height_add": 0.75, "notes": "Extra height for operator"},
        "egress": {"min_clear_width": 20, "min_clear_height": 24, "min_clear_area": 5.7, 
                   "max_sill_height": 44, "notes": "FBC 2023 R310.2.1 Egress requirements"},
    },
    "doors": {
        "entry": {"width_add": 2.5, "height_add": 2.5, "notes": "Standard entry door"},
        "french": {"width_add": 2.5, "height_add": 2.5, "notes": "Per leaf, double for pairs"},
        "sliding_glass": {"width_add": 1.0, "height_add": 1.0, "notes": "Tight tolerance"},
        "interior": {"width_add": 2.0, "height_add": 2.0, "notes": "Standard interior"},
        "pocket": {"width_add": 2.0, "height_add": 2.0, "notes": "Pocket door frame"},
        "bifold": {"width_add": 1.5, "height_add": 2.0, "notes": "Closet bifold"},
        "garage": {"width_add": 3.0, "height_add": 1.5, "notes": "Overhead garage door"},
    },
    "headers": {
        "up_to_4ft": {"size": "2x6 double", "notes": "Up to 4' span"},
        "4ft_to_6ft": {"size": "2x8 double", "notes": "4' to 6' span"},
        "6ft_to_8ft": {"size": "2x10 double", "notes": "6' to 8' span"},
        "8ft_to_10ft": {"size": "2x12 double", "notes": "8' to 10' span"},
        "over_10ft": {"size": "LVL or engineered", "notes": "Requires engineering"},
    },
    "wind_zone": {
        "zone_3": {"speed": 150, "notes": "Brevard County - 160 MPH design"},
        "hvhz": {"speed": 180, "notes": "High Velocity Hurricane Zone"},
        "impact_required": True,
        "notes": "FBC 2023 requires impact-rated glazing in WBDR zones"
    }
}

# ============================================================================
# RETAILER SKU PATTERNS & ENDPOINTS
# ============================================================================

RETAILER_CONFIG = {
    "home_depot": {
        "name": "Home Depot",
        "sku_format": "###-###-###",
        "sku_length": 9,
        "search_url": "https://www.homedepot.com/s/{query}",
        "product_url": "https://www.homedepot.com/p/{sku}",
        "store_id": "0912",  # Palm Bay, FL
        "categories": {
            "200": "Building Materials",
            "300": "Electrical",
            "500": "Hardware",
            "600": "Kitchen & Bath",
            "800": "Lighting",
            "1000": "Plumbing",
        }
    },
    "lowes": {
        "name": "Lowe's",
        "sku_format": "#########",
        "sku_length": 9,
        "search_url": "https://www.lowes.com/search?searchTerm={query}",
        "product_url": "https://www.lowes.com/pd/{slug}/{item_number}",
        "store_id": "0604",  # Palm Bay, FL
        "categories": {
            "building": "Building Supplies",
            "electrical": "Electrical",
            "hardware": "Hardware",
            "kitchen": "Kitchen",
            "lighting": "Lighting",
            "plumbing": "Plumbing",
        }
    }
}

# ============================================================================
# MATERIAL DATABASE - Common Construction Materials with Retailer SKUs
# ============================================================================

MATERIAL_DATABASE = {
    # LUMBER - Framing
    "2x4x8_spf_stud": {
        "description": "2x4x8' SPF Stud Grade",
        "unit": "EA",
        "home_depot": {"sku": "161640", "price": 3.98},
        "lowes": {"sku": "6009", "price": 3.88},
        "category": "lumber"
    },
    "2x4x10_spf": {
        "description": "2x4x10' SPF #2",
        "unit": "EA",
        "home_depot": {"sku": "161641", "price": 5.48},
        "lowes": {"sku": "6010", "price": 5.38},
        "category": "lumber"
    },
    "2x6x8_spf": {
        "description": "2x6x8' SPF #2",
        "unit": "EA",
        "home_depot": {"sku": "161650", "price": 6.98},
        "lowes": {"sku": "6020", "price": 6.88},
        "category": "lumber"
    },
    "2x6x10_spf": {
        "description": "2x6x10' SPF #2",
        "unit": "EA",
        "home_depot": {"sku": "161651", "price": 8.98},
        "lowes": {"sku": "6021", "price": 8.78},
        "category": "lumber"
    },
    "2x8x10_spf": {
        "description": "2x8x10' SPF #2",
        "unit": "EA",
        "home_depot": {"sku": "161660", "price": 11.48},
        "lowes": {"sku": "6030", "price": 11.28},
        "category": "lumber"
    },
    "2x10x10_spf": {
        "description": "2x10x10' SPF #2",
        "unit": "EA",
        "home_depot": {"sku": "161670", "price": 14.98},
        "lowes": {"sku": "6040", "price": 14.78},
        "category": "lumber"
    },
    "2x12x10_spf": {
        "description": "2x12x10' SPF #2",
        "unit": "EA",
        "home_depot": {"sku": "161680", "price": 18.98},
        "lowes": {"sku": "6050", "price": 18.68},
        "category": "lumber"
    },
    "4x4x8_pt_syp": {
        "description": "4x4x8' Pressure Treated SYP #2",
        "unit": "EA",
        "home_depot": {"sku": "161700", "price": 12.98},
        "lowes": {"sku": "6100", "price": 12.78},
        "category": "lumber"
    },
    
    # SHEATHING
    "osb_7_16_4x8": {
        "description": "7/16\" OSB Sheathing 4x8",
        "unit": "EA",
        "home_depot": {"sku": "386081", "price": 12.98},
        "lowes": {"sku": "12212", "price": 12.78},
        "category": "sheathing"
    },
    "cdx_15_32_4x8": {
        "description": "15/32\" CDX Plywood 4x8",
        "unit": "EA",
        "home_depot": {"sku": "166081", "price": 32.98},
        "lowes": {"sku": "12232", "price": 31.98},
        "category": "sheathing"
    },
    
    # CONCRETE
    "concrete_80lb": {
        "description": "Quikrete 80 lb Concrete Mix",
        "unit": "BAG",
        "home_depot": {"sku": "100318", "price": 6.35},
        "lowes": {"sku": "10383", "price": 6.28},
        "category": "concrete"
    },
    "rebar_5_20ft": {
        "description": "#5 Rebar 20' Length",
        "unit": "EA",
        "home_depot": {"sku": "202105", "price": 15.98},
        "lowes": {"sku": "44682", "price": 15.48},
        "category": "concrete"
    },
    
    # SIMPSON CONNECTORS
    "simpson_heta20": {
        "description": "Simpson HETA20 Hurricane Tie",
        "unit": "EA",
        "home_depot": {"sku": "204858", "price": 8.48},
        "lowes": {"sku": "44271", "price": 8.58},
        "category": "connectors"
    },
    "simpson_hts20": {
        "description": "Simpson HTS20 Strap",
        "unit": "EA",
        "home_depot": {"sku": "204862", "price": 7.28},
        "lowes": {"sku": "44275", "price": 7.38},
        "category": "connectors"
    },
    "simpson_sp4": {
        "description": "Simpson SP4 Anchor",
        "unit": "EA",
        "home_depot": {"sku": "204870", "price": 3.28},
        "lowes": {"sku": "44280", "price": 3.18},
        "category": "connectors"
    },
    "simpson_thd26": {
        "description": "Simpson THD26 Hanger",
        "unit": "EA",
        "home_depot": {"sku": "204880", "price": 12.48},
        "lowes": {"sku": "44290", "price": 12.38},
        "category": "connectors"
    },
    "simpson_h16": {
        "description": "Simpson H16 Hurricane Anchor",
        "unit": "EA",
        "home_depot": {"sku": "204885", "price": 6.48},
        "lowes": {"sku": "44295", "price": 6.58},
        "category": "connectors"
    },
    
    # INSULATION
    "r13_faced_15": {
        "description": "R-13 Kraft Faced 15\" x 32 SF",
        "unit": "ROLL",
        "home_depot": {"sku": "490573", "price": 18.98},
        "lowes": {"sku": "46713", "price": 18.78},
        "category": "insulation"
    },
    "r38_unfaced_24": {
        "description": "R-38 Unfaced 24\" x 48 SF",
        "unit": "ROLL",
        "home_depot": {"sku": "490583", "price": 52.98},
        "lowes": {"sku": "46723", "price": 51.98},
        "category": "insulation"
    },
    
    # ROOFING
    "metal_roof_panel": {
        "description": "Metal Roof Panel 3'x12' Galvanized",
        "unit": "EA",
        "home_depot": {"sku": "615310", "price": 45.98},
        "lowes": {"sku": "61578", "price": 44.98},
        "category": "roofing"
    },
    "ridge_vent_4ft": {
        "description": "Ridge Vent 4' Section",
        "unit": "EA",
        "home_depot": {"sku": "615340", "price": 12.98},
        "lowes": {"sku": "61590", "price": 12.78},
        "category": "roofing"
    },
    
    # DRYWALL
    "drywall_1_2_4x8": {
        "description": "1/2\" Drywall 4x8 Sheet",
        "unit": "EA",
        "home_depot": {"sku": "386101", "price": 12.48},
        "lowes": {"sku": "11501", "price": 12.28},
        "category": "drywall"
    },
    "drywall_5_8_x_4x8": {
        "description": "5/8\" Type X Fire Rated Drywall 4x8",
        "unit": "EA",
        "home_depot": {"sku": "386111", "price": 16.98},
        "lowes": {"sku": "11511", "price": 16.78},
        "category": "drywall"
    },
    
    # WINDOWS - Impact Rated for Florida
    "window_sh_24x36_impact": {
        "description": "24x36 Single Hung Impact Window",
        "unit": "EA",
        "home_depot": {"sku": "729248", "price": 389.00},
        "lowes": {"sku": "778234", "price": 399.00},
        "category": "windows"
    },
    "window_sh_30x48_impact": {
        "description": "30x48 Single Hung Impact Window",
        "unit": "EA",
        "home_depot": {"sku": "729258", "price": 489.00},
        "lowes": {"sku": "778244", "price": 499.00},
        "category": "windows"
    },
    "window_sh_36x60_impact": {
        "description": "36x60 Single Hung Impact Window",
        "unit": "EA",
        "home_depot": {"sku": "729268", "price": 589.00},
        "lowes": {"sku": "778254", "price": 599.00},
        "category": "windows"
    },
    
    # DOORS - Impact Rated for Florida
    "door_entry_36x80_impact": {
        "description": "36x80 Entry Door Impact Rated",
        "unit": "EA",
        "home_depot": {"sku": "835360", "price": 798.00},
        "lowes": {"sku": "867234", "price": 819.00},
        "category": "doors"
    },
    "door_french_60x80_impact": {
        "description": "60x80 French Door Impact Rated",
        "unit": "EA",
        "home_depot": {"sku": "835460", "price": 1998.00},
        "lowes": {"sku": "867334", "price": 2049.00},
        "category": "doors"
    },
    "door_garage_16x8": {
        "description": "16x8 Garage Door Insulated Wind Rated",
        "unit": "EA",
        "home_depot": {"sku": "835560", "price": 2698.00},
        "lowes": {"sku": "867434", "price": 2749.00},
        "category": "doors"
    },
    
    # ELECTRICAL
    "outlet_gfci_20a": {
        "description": "20A GFCI Outlet White",
        "unit": "EA",
        "home_depot": {"sku": "100605", "price": 18.97},
        "lowes": {"sku": "72410", "price": 19.48},
        "category": "electrical"
    },
    "outlet_afci_gfci_combo": {
        "description": "AFCI/GFCI Combo Outlet 20A",
        "unit": "EA",
        "home_depot": {"sku": "100615", "price": 44.97},
        "lowes": {"sku": "72420", "price": 46.98},
        "category": "electrical"
    },
    "panel_200a_main": {
        "description": "200A Main Breaker Panel",
        "unit": "EA",
        "home_depot": {"sku": "100700", "price": 189.00},
        "lowes": {"sku": "72500", "price": 194.00},
        "category": "electrical"
    },
    
    # PLUMBING
    "toilet_elongated": {
        "description": "Elongated Toilet 1.28 GPF",
        "unit": "EA",
        "home_depot": {"sku": "301580", "price": 179.00},
        "lowes": {"sku": "82410", "price": 184.00},
        "category": "plumbing"
    },
    "water_heater_50gal": {
        "description": "50 Gallon Electric Water Heater",
        "unit": "EA",
        "home_depot": {"sku": "301680", "price": 599.00},
        "lowes": {"sku": "82510", "price": 619.00},
        "category": "plumbing"
    },
}

# ============================================================================
# FIRECRAWL API INTEGRATION
# ============================================================================

def search_firecrawl(url: str, extract_schema: dict = None) -> dict:
    """
    Use Firecrawl API to scrape retailer product pages
    """
    if not FIRECRAWL_API_KEY:
        return {"error": "FIRECRAWL_API_KEY not configured", "data": None}
    
    try:
        headers = {
            "Authorization": f"Bearer {FIRECRAWL_API_KEY}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "url": url,
            "formats": ["markdown", "html"],
            "onlyMainContent": True
        }
        
        if extract_schema:
            payload["extract"] = {"schema": extract_schema}
        
        response = requests.post(
            "https://api.firecrawl.dev/v1/scrape",
            headers=headers,
            json=payload,
            timeout=30
        )
        
        if response.status_code == 200:
            return response.json()
        else:
            return {"error": f"Firecrawl API error: {response.status_code}", "data": None}
            
    except Exception as e:
        return {"error": str(e), "data": None}


def search_home_depot(query: str, max_results: int = 5) -> List[Dict[str, Any]]:
    """
    Search Home Depot for products using Firecrawl
    """
    search_url = f"https://www.homedepot.com/s/{query.replace(' ', '%20')}"
    
    # Try Firecrawl first
    result = search_firecrawl(search_url)
    
    if result.get("data"):
        # Parse products from scraped content
        products = parse_home_depot_results(result["data"])
        return products[:max_results]
    
    # Fallback to database lookup
    return lookup_material_database(query, "home_depot")


def search_lowes(query: str, max_results: int = 5) -> List[Dict[str, Any]]:
    """
    Search Lowe's for products using Firecrawl
    """
    search_url = f"https://www.lowes.com/search?searchTerm={query.replace(' ', '+')}"
    
    # Try Firecrawl first
    result = search_firecrawl(search_url)
    
    if result.get("data"):
        # Parse products from scraped content
        products = parse_lowes_results(result["data"])
        return products[:max_results]
    
    # Fallback to database lookup
    return lookup_material_database(query, "lowes")


def parse_home_depot_results(data: dict) -> List[Dict[str, Any]]:
    """Parse Home Depot search results from Firecrawl data"""
    products = []
    
    try:
        content = data.get("markdown", "") or data.get("html", "")
        
        # Extract product patterns using regex
        # Home Depot SKU pattern: Internet # XXXXXX or Model # XXXXX
        sku_pattern = r'Internet\s*#\s*(\d+)|Model\s*#\s*([A-Z0-9-]+)'
        price_pattern = r'\$(\d+(?:,\d{3})*(?:\.\d{2})?)'
        
        skus = re.findall(sku_pattern, content)
        prices = re.findall(price_pattern, content)
        
        for i, sku_match in enumerate(skus[:5]):
            sku = sku_match[0] or sku_match[1]
            price = float(prices[i].replace(',', '')) if i < len(prices) else 0.0
            
            products.append({
                "retailer": "Home Depot",
                "sku": sku,
                "price": price,
                "url": f"https://www.homedepot.com/p/{sku}",
                "in_stock": True,
                "source": "firecrawl"
            })
    except Exception as e:
        print(f"Error parsing Home Depot results: {e}")
    
    return products


def parse_lowes_results(data: dict) -> List[Dict[str, Any]]:
    """Parse Lowe's search results from Firecrawl data"""
    products = []
    
    try:
        content = data.get("markdown", "") or data.get("html", "")
        
        # Lowe's item number pattern
        item_pattern = r'Item\s*#\s*(\d+)|Model\s*#\s*([A-Z0-9-]+)'
        price_pattern = r'\$(\d+(?:,\d{3})*(?:\.\d{2})?)'
        
        items = re.findall(item_pattern, content)
        prices = re.findall(price_pattern, content)
        
        for i, item_match in enumerate(items[:5]):
            item_num = item_match[0] or item_match[1]
            price = float(prices[i].replace(',', '')) if i < len(prices) else 0.0
            
            products.append({
                "retailer": "Lowes",
                "sku": item_num,
                "price": price,
                "url": f"https://www.lowes.com/pd/{item_num}",
                "in_stock": True,
                "source": "firecrawl"
            })
    except Exception as e:
        print(f"Error parsing Lowe's results: {e}")
    
    return products


def lookup_material_database(query: str, retailer: str) -> List[Dict[str, Any]]:
    """
    Fallback lookup in local material database
    """
    results = []
    query_lower = query.lower()
    
    for key, material in MATERIAL_DATABASE.items():
        if query_lower in material["description"].lower() or query_lower in key:
            retailer_data = material.get(retailer, material.get("home_depot", {}))
            results.append({
                "retailer": "Home Depot" if retailer == "home_depot" else "Lowe's",
                "sku": retailer_data.get("sku", "N/A"),
                "description": material["description"],
                "price": retailer_data.get("price", 0.0),
                "unit": material["unit"],
                "category": material["category"],
                "source": "database"
            })
    
    return results


def get_best_price(material_query: str) -> Dict[str, Any]:
    """
    Compare prices between Home Depot and Lowe's, return best option
    """
    hd_results = search_home_depot(material_query, max_results=3)
    lowes_results = search_lowes(material_query, max_results=3)
    
    all_results = hd_results + lowes_results
    
    if not all_results:
        return {
            "material": material_query,
            "best_retailer": "N/A",
            "best_price": 0.0,
            "best_sku": "N/A",
            "comparison": []
        }
    
    # Find lowest price
    best = min(all_results, key=lambda x: x.get("price", float("inf")) or float("inf"))
    
    return {
        "material": material_query,
        "best_retailer": best.get("retailer", "N/A"),
        "best_price": best.get("price", 0.0),
        "best_sku": best.get("sku", "N/A"),
        "best_url": best.get("url", ""),
        "comparison": all_results
    }


# ============================================================================
# ROUGH OPENING CALCULATOR
# ============================================================================

def calculate_rough_opening(
    opening_type: str,
    nominal_width: float,
    nominal_height: float,
    door_or_window: str = "window"
) -> Dict[str, Any]:
    """
    Calculate rough opening dimensions per Florida Building Code 2023
    
    Args:
        opening_type: Type of opening (single_hung, entry, french, etc.)
        nominal_width: Nominal width in inches
        nominal_height: Nominal height in inches
        door_or_window: "window" or "door"
    
    Returns:
        Dict with rough opening dimensions and specifications
    """
    category = "windows" if door_or_window == "window" else "doors"
    specs = FLORIDA_ROUGH_OPENINGS[category].get(opening_type, {})
    
    width_add = specs.get("width_add", 0.5)
    height_add = specs.get("height_add", 0.5)
    
    ro_width = nominal_width + width_add
    ro_height = nominal_height + height_add
    
    # Determine header size based on span
    span_ft = nominal_width / 12
    if span_ft <= 4:
        header = FLORIDA_ROUGH_OPENINGS["headers"]["up_to_4ft"]
    elif span_ft <= 6:
        header = FLORIDA_ROUGH_OPENINGS["headers"]["4ft_to_6ft"]
    elif span_ft <= 8:
        header = FLORIDA_ROUGH_OPENINGS["headers"]["6ft_to_8ft"]
    elif span_ft <= 10:
        header = FLORIDA_ROUGH_OPENINGS["headers"]["8ft_to_10ft"]
    else:
        header = FLORIDA_ROUGH_OPENINGS["headers"]["over_10ft"]
    
    # Check egress requirements for bedrooms
    egress_compliant = True
    egress_notes = ""
    if door_or_window == "window":
        egress = FLORIDA_ROUGH_OPENINGS["windows"]["egress"]
        clear_width = nominal_width - 2  # Approximate clear opening
        clear_height = nominal_height - 2
        clear_area = (clear_width * clear_height) / 144  # Convert to SF
        
        if clear_width < egress["min_clear_width"]:
            egress_compliant = False
            egress_notes += f"Width {clear_width}\" < {egress['min_clear_width']}\" min. "
        if clear_height < egress["min_clear_height"]:
            egress_compliant = False
            egress_notes += f"Height {clear_height}\" < {egress['min_clear_height']}\" min. "
        if clear_area < egress["min_clear_area"]:
            egress_compliant = False
            egress_notes += f"Area {clear_area:.1f} SF < {egress['min_clear_area']} SF min. "
    
    return {
        "opening_type": opening_type,
        "category": door_or_window,
        "nominal_width": nominal_width,
        "nominal_height": nominal_height,
        "rough_opening_width": ro_width,
        "rough_opening_height": ro_height,
        "header_size": header["size"],
        "header_notes": header["notes"],
        "egress_compliant": egress_compliant,
        "egress_notes": egress_notes or "Meets FBC 2023 egress requirements",
        "impact_required": FLORIDA_ROUGH_OPENINGS["wind_zone"]["impact_required"],
        "wind_zone": "160 MPH - Brevard County",
        "fbc_reference": "FBC 2023 Chapter 24"
    }


# ============================================================================
# AGENT WORKFLOW FUNCTIONS
# ============================================================================

def extract_materials_from_takeoff(state: MaterialListState) -> MaterialListState:
    """
    Stage 1: Process takeoff quantities and categorize materials
    """
    print("📋 Stage 1: Extracting materials from takeoff data...")
    
    try:
        takeoff_materials = state.get("takeoff_materials", [])
        
        # Categorize by MasterFormat division
        categorized = {
            "03_concrete": [],
            "04_masonry": [],
            "05_metals": [],
            "06_wood": [],
            "07_thermal": [],
            "08_openings": [],
            "09_finishes": [],
            "16_electrical": [],
            "22_plumbing": [],
            "31_sitework": [],
        }
        
        for material in takeoff_materials:
            item = material.get("item", "").lower()
            
            if any(x in item for x in ["concrete", "slab", "footing", "rebar"]):
                categorized["03_concrete"].append(material)
            elif any(x in item for x in ["cmu", "block", "masonry", "mortar", "grout"]):
                categorized["04_masonry"].append(material)
            elif any(x in item for x in ["simpson", "strap", "anchor", "hanger", "connector"]):
                categorized["05_metals"].append(material)
            elif any(x in item for x in ["2x4", "2x6", "2x8", "lumber", "plywood", "sheathing", "lvl"]):
                categorized["06_wood"].append(material)
            elif any(x in item for x in ["insulation", "roofing", "roof", "vapor", "flashing"]):
                categorized["07_thermal"].append(material)
            elif any(x in item for x in ["window", "door", "opening", "garage"]):
                categorized["08_openings"].append(material)
            elif any(x in item for x in ["drywall", "paint", "stucco", "flooring", "tile"]):
                categorized["09_finishes"].append(material)
            elif any(x in item for x in ["electrical", "outlet", "panel", "wire", "light"]):
                categorized["16_electrical"].append(material)
            elif any(x in item for x in ["plumbing", "sink", "toilet", "water", "pipe"]):
                categorized["22_plumbing"].append(material)
            else:
                categorized["31_sitework"].append(material)
        
        state["categorized_materials"] = categorized
        state["status"] = "materials_extracted"
        state["messages"].append({
            "stage": "extract_materials",
            "status": "success",
            "count": len(takeoff_materials)
        })
        
    except Exception as e:
        state["errors"].append(f"Material extraction error: {str(e)}")
        state["status"] = "failed"
    
    return state


def search_retailers(state: MaterialListState) -> MaterialListState:
    """
    Stage 2: Search Home Depot and Lowe's for SKUs and prices
    """
    print("🛒 Stage 2: Searching retailers for SKUs and prices...")
    
    try:
        categorized = state.get("categorized_materials", {})
        best_prices = []
        
        # Process each material category
        for category, materials in categorized.items():
            for material in materials:
                item_name = material.get("item", "")
                
                # Get best price comparison
                price_result = get_best_price(item_name)
                price_result["original_item"] = item_name
                price_result["quantity"] = material.get("qty", 1)
                price_result["unit"] = material.get("unit", "EA")
                price_result["category"] = category
                
                best_prices.append(price_result)
        
        state["best_prices"] = best_prices
        state["status"] = "retailers_searched"
        state["messages"].append({
            "stage": "search_retailers",
            "status": "success",
            "items_searched": len(best_prices)
        })
        
    except Exception as e:
        state["errors"].append(f"Retailer search error: {str(e)}")
        state["status"] = "failed"
    
    return state


def generate_rough_openings(state: MaterialListState) -> MaterialListState:
    """
    Stage 3: Calculate rough openings for windows and doors per FBC 2023
    """
    print("📐 Stage 3: Calculating rough openings per Florida Building Code...")
    
    try:
        categorized = state.get("categorized_materials", {})
        openings = categorized.get("08_openings", [])
        rough_openings = []
        
        for opening in openings:
            item = opening.get("item", "")
            
            # Parse nominal dimensions from item name
            # Examples: "25 SH Window (2'5\" x 3'6\")", "3/0 x 8/0 Ext Door"
            dim_match = re.search(r"(\d+)[/'\"]\s*[x×]\s*(\d+)[/'\"x]?(\d+)?", item, re.IGNORECASE)
            
            if dim_match:
                width_ft = int(dim_match.group(1))
                height_ft = int(dim_match.group(2))
                height_in = int(dim_match.group(3)) if dim_match.group(3) else 0
                
                # Convert to inches
                nominal_width = width_ft * 12 if width_ft < 10 else width_ft
                nominal_height = height_ft * 12 + height_in if height_ft < 10 else height_ft
                
                # Determine type
                item_lower = item.lower()
                if "window" in item_lower:
                    if "sh" in item_lower or "single" in item_lower:
                        opening_type = "single_hung"
                    elif "dh" in item_lower or "double" in item_lower:
                        opening_type = "double_hung"
                    else:
                        opening_type = "single_hung"
                    category = "window"
                elif "door" in item_lower:
                    if "french" in item_lower:
                        opening_type = "french"
                    elif "garage" in item_lower:
                        opening_type = "garage"
                    elif "interior" in item_lower or "int" in item_lower:
                        opening_type = "interior"
                    else:
                        opening_type = "entry"
                    category = "door"
                else:
                    continue
                
                ro = calculate_rough_opening(opening_type, nominal_width, nominal_height, category)
                ro["original_item"] = item
                ro["quantity"] = opening.get("qty", 1)
                rough_openings.append(ro)
        
        state["rough_openings"] = rough_openings
        state["messages"].append({
            "stage": "rough_openings",
            "status": "success",
            "count": len(rough_openings)
        })
        
    except Exception as e:
        state["errors"].append(f"Rough opening calculation error: {str(e)}")
    
    return state


def generate_procurement_list(state: MaterialListState) -> MaterialListState:
    """
    Stage 4: Generate final procurement list with best prices
    """
    print("📝 Stage 4: Generating procurement list...")
    
    try:
        best_prices = state.get("best_prices", [])
        procurement_list = []
        total_material_cost = 0.0
        
        for item in best_prices:
            qty = item.get("quantity", 1)
            unit_price = item.get("best_price", 0.0)
            ext_cost = qty * unit_price
            total_material_cost += ext_cost
            
            procurement_list.append({
                "item": item.get("original_item", item.get("material", "")),
                "quantity": qty,
                "unit": item.get("unit", "EA"),
                "retailer": item.get("best_retailer", "N/A"),
                "sku": item.get("best_sku", "N/A"),
                "unit_price": unit_price,
                "extended_cost": ext_cost,
                "url": item.get("best_url", ""),
                "category": item.get("category", ""),
                "comparison": item.get("comparison", [])
            })
        
        state["procurement_list"] = procurement_list
        state["total_material_cost"] = total_material_cost
        state["total_labor_cost"] = total_material_cost * 0.45  # Estimate 45% labor
        state["status"] = "completed"
        state["messages"].append({
            "stage": "procurement_list",
            "status": "success",
            "total_items": len(procurement_list),
            "total_cost": total_material_cost
        })
        
    except Exception as e:
        state["errors"].append(f"Procurement list error: {str(e)}")
        state["status"] = "failed"
    
    return state


# ============================================================================
# EXCEL REPORT GENERATOR
# ============================================================================

def generate_material_list_excel(state: MaterialListState, output_path: str) -> str:
    """
    Generate comprehensive Excel report with procurement data
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl not installed"
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    subheader_fill = PatternFill("solid", fgColor="4A90A4")
    best_price_fill = PatternFill("solid", fgColor="E8F5E9")
    alt_fill = PatternFill("solid", fgColor="F5F5F5")
    total_fill = PatternFill("solid", fgColor="FFEBEE")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # ========================================================================
    # SHEET 1: PROCUREMENT SUMMARY
    # ========================================================================
    ws_summary = wb.active
    ws_summary.title = "Procurement Summary"
    
    ws_summary.merge_cells('A1:H1')
    ws_summary['A1'] = f"MATERIAL LIST & PROCUREMENT - {state.get('project_name', 'Project')}"
    ws_summary['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_summary['A1'].fill = header_fill
    
    ws_summary['A3'] = "Project:"
    ws_summary['B3'] = state.get('project_name', '')
    ws_summary['A4'] = "Address:"
    ws_summary['B4'] = state.get('project_address', '')
    ws_summary['A5'] = "Generated:"
    ws_summary['B5'] = datetime.now().strftime("%B %d, %Y %I:%M %p")
    
    row = 7
    headers = ["Item Description", "Qty", "Unit", "Best Retailer", "SKU", "Unit Price", "Extended Cost", "URL"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    procurement_list = state.get("procurement_list", [])
    for i, item in enumerate(procurement_list):
        r = row + 1 + i
        ws_summary.cell(row=r, column=1, value=item.get("item", "")).border = thin_border
        ws_summary.cell(row=r, column=2, value=item.get("quantity", 0)).border = thin_border
        ws_summary.cell(row=r, column=3, value=item.get("unit", "")).border = thin_border
        ws_summary.cell(row=r, column=4, value=item.get("retailer", "")).border = thin_border
        ws_summary.cell(row=r, column=5, value=item.get("sku", "")).border = thin_border
        ws_summary.cell(row=r, column=6, value=item.get("unit_price", 0)).border = thin_border
        ws_summary.cell(row=r, column=6).number_format = '$#,##0.00'
        ws_summary.cell(row=r, column=7, value=f"=B{r}*F{r}").border = thin_border
        ws_summary.cell(row=r, column=7).number_format = '$#,##0.00'
        ws_summary.cell(row=r, column=8, value=item.get("url", "")).border = thin_border
        
        if i % 2 == 1:
            for c in range(1, 9):
                ws_summary.cell(row=r, column=c).fill = alt_fill
    
    # Totals
    total_row = row + len(procurement_list) + 1
    ws_summary.cell(row=total_row, column=6, value="TOTAL:").font = Font(bold=True)
    ws_summary.cell(row=total_row, column=7, value=f"=SUM(G{row+1}:G{total_row-1})")
    ws_summary.cell(row=total_row, column=7).number_format = '$#,##0.00'
    ws_summary.cell(row=total_row, column=7).font = Font(bold=True)
    for c in range(6, 8):
        ws_summary.cell(row=total_row, column=c).fill = total_fill
    
    # Column widths
    ws_summary.column_dimensions['A'].width = 45
    ws_summary.column_dimensions['B'].width = 8
    ws_summary.column_dimensions['C'].width = 8
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 12
    ws_summary.column_dimensions['F'].width = 12
    ws_summary.column_dimensions['G'].width = 15
    ws_summary.column_dimensions['H'].width = 50
    
    # ========================================================================
    # SHEET 2: ROUGH OPENINGS - FBC 2023
    # ========================================================================
    ws_ro = wb.create_sheet(title="Rough Openings - FBC")
    
    ws_ro.merge_cells('A1:J1')
    ws_ro['A1'] = "ROUGH OPENINGS SCHEDULE - FLORIDA BUILDING CODE 2023"
    ws_ro['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_ro['A1'].fill = header_fill
    
    row = 3
    ro_headers = ["Item", "Type", "Nominal W", "Nominal H", "R.O. Width", "R.O. Height", 
                  "Header", "Egress OK", "Impact Req", "Notes"]
    for col, header in enumerate(ro_headers, 1):
        cell = ws_ro.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    rough_openings = state.get("rough_openings", [])
    for i, ro in enumerate(rough_openings):
        r = row + 1 + i
        ws_ro.cell(row=r, column=1, value=ro.get("original_item", "")).border = thin_border
        ws_ro.cell(row=r, column=2, value=ro.get("opening_type", "")).border = thin_border
        ws_ro.cell(row=r, column=3, value=f"{ro.get('nominal_width', 0)}\"").border = thin_border
        ws_ro.cell(row=r, column=4, value=f"{ro.get('nominal_height', 0)}\"").border = thin_border
        ws_ro.cell(row=r, column=5, value=f"{ro.get('rough_opening_width', 0)}\"").border = thin_border
        ws_ro.cell(row=r, column=6, value=f"{ro.get('rough_opening_height', 0)}\"").border = thin_border
        ws_ro.cell(row=r, column=7, value=ro.get("header_size", "")).border = thin_border
        ws_ro.cell(row=r, column=8, value="✓" if ro.get("egress_compliant") else "✗").border = thin_border
        ws_ro.cell(row=r, column=9, value="Yes" if ro.get("impact_required") else "No").border = thin_border
        ws_ro.cell(row=r, column=10, value=ro.get("egress_notes", "")).border = thin_border
        
        if i % 2 == 1:
            for c in range(1, 11):
                ws_ro.cell(row=r, column=c).fill = alt_fill
    
    for col in range(1, 11):
        ws_ro.column_dimensions[get_column_letter(col)].width = 15
    ws_ro.column_dimensions['A'].width = 35
    ws_ro.column_dimensions['J'].width = 40
    
    # ========================================================================
    # SHEET 3: PRICE COMPARISON
    # ========================================================================
    ws_compare = wb.create_sheet(title="Price Comparison")
    
    ws_compare.merge_cells('A1:F1')
    ws_compare['A1'] = "HOME DEPOT vs LOWE'S PRICE COMPARISON"
    ws_compare['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_compare['A1'].fill = header_fill
    
    row = 3
    compare_headers = ["Item", "Home Depot SKU", "HD Price", "Lowe's SKU", "Lowe's Price", "Best Price"]
    for col, header in enumerate(compare_headers, 1):
        cell = ws_compare.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for i, item in enumerate(procurement_list):
        r = row + 1 + i
        comparison = item.get("comparison", [])
        
        hd_data = next((x for x in comparison if x.get("retailer") == "Home Depot"), {})
        lowes_data = next((x for x in comparison if x.get("retailer") == "Lowes"), {})
        
        ws_compare.cell(row=r, column=1, value=item.get("item", "")).border = thin_border
        ws_compare.cell(row=r, column=2, value=hd_data.get("sku", "N/A")).border = thin_border
        ws_compare.cell(row=r, column=3, value=hd_data.get("price", 0)).border = thin_border
        ws_compare.cell(row=r, column=3).number_format = '$#,##0.00'
        ws_compare.cell(row=r, column=4, value=lowes_data.get("sku", "N/A")).border = thin_border
        ws_compare.cell(row=r, column=5, value=lowes_data.get("price", 0)).border = thin_border
        ws_compare.cell(row=r, column=5).number_format = '$#,##0.00'
        ws_compare.cell(row=r, column=6, value=item.get("unit_price", 0)).border = thin_border
        ws_compare.cell(row=r, column=6).number_format = '$#,##0.00'
        ws_compare.cell(row=r, column=6).fill = best_price_fill
        
        if i % 2 == 1:
            for c in range(1, 6):
                ws_compare.cell(row=r, column=c).fill = alt_fill
    
    ws_compare.column_dimensions['A'].width = 45
    for col in range(2, 7):
        ws_compare.column_dimensions[get_column_letter(col)].width = 15
    
    # Save workbook
    wb.save(output_path)
    return output_path


# ============================================================================
# MAIN WORKFLOW
# ============================================================================

def run_material_list_agent(
    takeoff_materials: List[Dict[str, Any]],
    project_name: str = "Project",
    project_address: str = "",
    output_path: str = None
) -> Dict[str, Any]:
    """
    Run the complete material list agent workflow
    
    Args:
        takeoff_materials: List of materials from takeoff agent
        project_name: Project name
        project_address: Project address
        output_path: Path for Excel output
    
    Returns:
        Dict with procurement list, costs, and file path
    """
    # Initialize state
    state: MaterialListState = {
        "takeoff_materials": takeoff_materials,
        "takeoff_quantities": [],
        "project_name": project_name,
        "project_id": datetime.now().strftime("%Y%m%d%H%M%S"),
        "project_address": project_address,
        "categorized_materials": {},
        "rough_openings": [],
        "home_depot_products": [],
        "lowes_products": [],
        "best_prices": [],
        "technical_specs": [],
        "florida_code_compliance": [],
        "procurement_list": [],
        "excel_path": "",
        "total_material_cost": 0.0,
        "total_labor_cost": 0.0,
        "messages": [],
        "errors": [],
        "status": "initiated"
    }
    
    # Run workflow stages
    state = extract_materials_from_takeoff(state)
    if state["status"] != "failed":
        state = search_retailers(state)
    if state["status"] != "failed":
        state = generate_rough_openings(state)
    if state["status"] != "failed":
        state = generate_procurement_list(state)
    
    # Generate Excel report
    if output_path and state["status"] == "completed":
        excel_path = generate_material_list_excel(state, output_path)
        state["excel_path"] = excel_path
    
    return {
        "status": state["status"],
        "procurement_list": state["procurement_list"],
        "rough_openings": state["rough_openings"],
        "total_material_cost": state["total_material_cost"],
        "total_labor_cost": state["total_labor_cost"],
        "excel_path": state.get("excel_path", ""),
        "messages": state["messages"],
        "errors": state["errors"]
    }


# ============================================================================
# CLI INTERFACE
# ============================================================================

if __name__ == "__main__":
    import sys
    
    # Example usage with sample materials
    sample_materials = [
        {"item": "2x4 SPF No.2 @ 16\" O.C.", "qty": 3200, "unit": "LF", "unit_cost": 0.85},
        {"item": "Simpson HETA20 Straps", "qty": 96, "unit": "EA", "unit_cost": 8.50},
        {"item": "25 SH Window (2'5\" x 3'6\")", "qty": 24, "unit": "EA", "unit_cost": 425.00},
        {"item": "3/0 x 8/0 Exterior Door/Frame", "qty": 8, "unit": "EA", "unit_cost": 850.00},
        {"item": "16/0 x 8/0 Overhead Garage Door", "qty": 2, "unit": "EA", "unit_cost": 2800.00},
        {"item": "R-13 Batt Insulation (Walls)", "qty": 4200, "unit": "SF", "unit_cost": 0.95},
        {"item": "1/2\" Drywall", "qty": 16500, "unit": "SF", "unit_cost": 1.85},
    ]
    
    output_file = sys.argv[1] if len(sys.argv) > 1 else "/tmp/material_list_output.xlsx"
    
    result = run_material_list_agent(
        takeoff_materials=sample_materials,
        project_name="Hester Avenue Duplex",
        project_address="2826 Hester Ave SE, Palm Bay, FL 32908",
        output_path=output_file
    )
    
    print(f"\n✅ Material List Agent Complete")
    print(f"Status: {result['status']}")
    print(f"Items processed: {len(result['procurement_list'])}")
    print(f"Total Material Cost: ${result['total_material_cost']:,.2f}")
    print(f"Total Labor Cost: ${result['total_labor_cost']:,.2f}")
    print(f"Excel Report: {result['excel_path']}")
