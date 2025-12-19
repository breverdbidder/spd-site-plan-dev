"""
Construction Takeoff Agent for SPD LangGraph Orchestration
===========================================================

Agentic AI system for automated construction quantity surveying from PDF plans.
Integrates with API mega library, Firecrawl for competitive intelligence, and 
generates Excel-based takeoff reports.

Author: Claude Sonnet 4.5 (AI Architect)
Date: December 19, 2025
Repository: github.com/breverdbidder/spd-site-plan-dev
Stack: LangGraph + Anthropic Claude + Firecrawl + OpenPyXL
"""

import os
import json
import base64
from pathlib import Path
from typing import TypedDict, Literal, Annotated, List, Dict, Any
from datetime import datetime
from langgraph.graph import StateGraph, END
from anthropic import Anthropic
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Initialize clients
anthropic_client = Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
FIRECRAWL_API_KEY = os.environ.get("FIRECRAWL_API_KEY", "")
SUPABASE_URL = "https://mocerqjnksmhcjzxrewo.supabase.co"
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

# ============================================================================
# STATE DEFINITION
# ============================================================================

class TakeoffState(TypedDict):
    """State for construction takeoff workflow"""
    # Input
    pdf_path: str
    pdf_base64: str
    project_name: str
    project_id: str
    
    # PDF Analysis
    pdf_extracted_text: str
    pdf_pages_analyzed: int
    drawing_type: str  # "site_plan", "floor_plan", "elevation", "detail"
    scale_detected: str
    
    # Quantity Extraction
    quantities: List[Dict[str, Any]]
    materials: List[Dict[str, Any]]
    measurements: Dict[str, Any]
    
    # Competitor Analysis
    competitors_analyzed: List[str]
    competitor_features: Dict[str, List[str]]
    best_practices: List[str]
    
    # Cost Estimation
    cost_estimates: Dict[str, Any]
    labor_hours: Dict[str, float]
    total_estimated_cost: float
    
    # Output
    excel_path: str
    report_summary: str
    
    # Metadata
    messages: List[Dict[str, str]]
    errors: List[str]
    status: Literal["initiated", "pdf_processed", "quantities_extracted", 
                    "competitors_analyzed", "costs_calculated", "report_generated", 
                    "completed", "failed"]

# ============================================================================
# COMPETITOR TARGETS - Best-in-class takeoff solutions
# ============================================================================

COMPETITOR_TARGETS = [
    {
        "name": "PlanSwift",
        "url": "https://www.planswift.com",
        "focus": "Digital takeoff and estimating software"
    },
    {
        "name": "Bluebeam Revu",
        "url": "https://www.bluebeam.com/solutions/takeoff",
        "focus": "PDF markup and quantity takeoff"
    },
    {
        "name": "On-Screen Takeoff",
        "url": "https://www.on-screen-takeoff.com",
        "focus": "Construction takeoff software"
    },
    {
        "name": "Stack Takeoff",
        "url": "https://www.stackct.com",
        "focus": "Cloud-based construction takeoff"
    },
    {
        "name": "STACK Construction Technologies",
        "url": "https://www.stackct.com/products/takeoff",
        "focus": "Preconstruction takeoff and estimating"
    }
]

# ============================================================================
# CONSTRUCTION MATERIAL CATEGORIES
# ============================================================================

MATERIAL_CATEGORIES = {
    "earthwork": ["excavation", "grading", "fill", "compaction", "topsoil"],
    "concrete": ["footings", "slabs", "walls", "columns", "beams", "piers"],
    "masonry": ["brick", "block", "stone", "mortar"],
    "metals": ["structural steel", "rebar", "metal studs", "anchors"],
    "wood_plastics": ["framing lumber", "plywood", "joists", "decking"],
    "thermal_moisture": ["insulation", "roofing", "waterproofing", "flashing"],
    "openings": ["doors", "windows", "frames", "hardware"],
    "finishes": ["drywall", "paint", "flooring", "tile", "carpet"],
    "specialties": ["signage", "lockers", "toilet partitions"],
    "equipment": ["appliances", "furniture", "fixtures"],
    "furnishings": ["cabinets", "countertops", "window treatments"],
    "mechanical": ["HVAC", "plumbing", "piping"],
    "electrical": ["wiring", "fixtures", "panels", "conduit"],
    "site_work": ["paving", "curbs", "landscaping", "utilities", "fencing"]
}

# ============================================================================
# UNIT COST DATABASE (Florida Market - 2025)
# ============================================================================

UNIT_COSTS = {
    # Earthwork (per cubic yard)
    "excavation": {"unit": "CY", "cost": 12.50, "labor_hours": 0.05},
    "grading": {"unit": "CY", "cost": 8.75, "labor_hours": 0.03},
    "fill": {"unit": "CY", "cost": 15.00, "labor_hours": 0.04},
    
    # Concrete (per cubic yard)
    "concrete_slab": {"unit": "CY", "cost": 150.00, "labor_hours": 0.75},
    "concrete_footing": {"unit": "CY", "cost": 175.00, "labor_hours": 1.00},
    "concrete_wall": {"unit": "CY", "cost": 200.00, "labor_hours": 1.25},
    
    # Masonry (per square foot)
    "brick_veneer": {"unit": "SF", "cost": 12.50, "labor_hours": 0.20},
    "concrete_block": {"unit": "SF", "cost": 8.75, "labor_hours": 0.15},
    
    # Framing (per board foot)
    "wood_framing": {"unit": "BF", "cost": 2.50, "labor_hours": 0.05},
    
    # Roofing (per square foot)
    "asphalt_shingle": {"unit": "SF", "cost": 3.50, "labor_hours": 0.08},
    "metal_roofing": {"unit": "SF", "cost": 8.00, "labor_hours": 0.12},
    
    # Finishes (per square foot)
    "drywall": {"unit": "SF", "cost": 2.25, "labor_hours": 0.06},
    "paint": {"unit": "SF", "cost": 1.50, "labor_hours": 0.04},
    "tile": {"unit": "SF", "cost": 8.00, "labor_hours": 0.15},
    
    # Site work (per linear foot)
    "paving": {"unit": "SF", "cost": 4.50, "labor_hours": 0.03},
    "curb_gutter": {"unit": "LF", "cost": 18.00, "labor_hours": 0.10},
    "fencing": {"unit": "LF", "cost": 25.00, "labor_hours": 0.15},
    
    # Default fallback
    "default": {"unit": "EA", "cost": 100.00, "labor_hours": 0.50}
}

# ============================================================================
# AGENT FUNCTIONS
# ============================================================================

def process_pdf(state: TakeoffState) -> TakeoffState:
    """
    Stage 1: Process PDF construction plans using Claude's vision capabilities
    """
    print("📄 Stage 1: Processing PDF construction plans...")
    
    try:
        pdf_path = state.get("pdf_path", "")
        
        if not pdf_path or not Path(pdf_path).exists():
            state["errors"].append("PDF file not found")
            state["status"] = "failed"
            return state
        
        # Read and encode PDF
        with open(pdf_path, "rb") as f:
            pdf_bytes = f.read()
            pdf_base64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")
        
        state["pdf_base64"] = pdf_base64
        
        # Analyze PDF with Claude
        message = anthropic_client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4000,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": pdf_base64
                        }
                    },
                    {
                        "type": "text",
                        "text": """Analyze this construction plan PDF and extract:

1. DOCUMENT TYPE: Identify if this is a site plan, floor plan, elevation, detail drawing, etc.
2. SCALE: Extract the drawing scale (e.g., 1/4" = 1'-0", 1" = 20', etc.)
3. DIMENSIONS: All visible dimensions and measurements
4. MATERIALS: List all specified materials with quantities
5. QUANTITIES: Count discrete items (doors, windows, fixtures, etc.)
6. AREAS: Calculate all areas shown (rooms, total building, site area, etc.)
7. LINEAR MEASUREMENTS: Total linear feet of walls, fencing, utilities, etc.
8. NOTES: Extract any specification notes or construction details

Format your response as structured JSON with these sections.
Focus on quantities that can be used for cost estimation."""
                    }
                ]
            }]
        )
        
        response_text = message.content[0].text
        state["pdf_extracted_text"] = response_text
        state["pdf_pages_analyzed"] = 1  # Would iterate for multi-page PDFs
        
        # Extract structured data from Claude's response
        try:
            # Attempt to parse JSON if Claude returned it
            import re
            json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
            if json_match:
                extracted_data = json.loads(json_match.group())
                state["drawing_type"] = extracted_data.get("document_type", "unknown")
                state["scale_detected"] = extracted_data.get("scale", "not specified")
            else:
                state["drawing_type"] = "site_plan"
                state["scale_detected"] = "not detected"
        except:
            state["drawing_type"] = "unknown"
            state["scale_detected"] = "not detected"
        
        state["status"] = "pdf_processed"
        state["messages"].append({
            "timestamp": datetime.utcnow().isoformat(),
            "stage": "pdf_processing",
            "message": f"PDF analyzed successfully. Type: {state['drawing_type']}"
        })
        
        print(f"✅ PDF processed: {state['drawing_type']}, Scale: {state['scale_detected']}")
        
    except Exception as e:
        error_msg = f"PDF processing failed: {str(e)}"
        state["errors"].append(error_msg)
        state["status"] = "failed"
        print(f"❌ {error_msg}")
    
    return state


def extract_quantities(state: TakeoffState) -> TakeoffState:
    """
    Stage 2: Extract quantities, materials, and measurements using AI
    """
    print("📊 Stage 2: Extracting quantities and materials...")
    
    try:
        # Use Claude to structure the extracted text into quantities
        message = anthropic_client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4000,
            messages=[{
                "role": "user",
                "content": f"""Based on this construction plan analysis, create a structured takeoff:

PLAN ANALYSIS:
{state['pdf_extracted_text']}

Generate a JSON response with:
1. quantities: Array of {{item, quantity, unit, category, notes}}
2. materials: Array of {{material, quantity, unit, category, specification}}
3. measurements: Object with {{total_area_sf, total_perimeter_lf, total_volume_cy}}

Use these categories: {', '.join(MATERIAL_CATEGORIES.keys())}

Focus on construction materials and quantities suitable for cost estimation.
If a measurement isn't specified, use "TBD" but include the item."""
            }]
        )
        
        response_text = message.content[0].text
        
        # Parse structured data
        import re
        json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
        if json_match:
            takeoff_data = json.loads(json_match.group())
            state["quantities"] = takeoff_data.get("quantities", [])
            state["materials"] = takeoff_data.get("materials", [])
            state["measurements"] = takeoff_data.get("measurements", {})
        else:
            # Fallback: create basic structure
            state["quantities"] = [
                {"item": "Excavation", "quantity": "TBD", "unit": "CY", "category": "earthwork"},
                {"item": "Concrete Slab", "quantity": "TBD", "unit": "CY", "category": "concrete"}
            ]
            state["materials"] = []
            state["measurements"] = {}
        
        state["status"] = "quantities_extracted"
        state["messages"].append({
            "timestamp": datetime.utcnow().isoformat(),
            "stage": "quantity_extraction",
            "message": f"Extracted {len(state['quantities'])} quantities, {len(state['materials'])} materials"
        })
        
        print(f"✅ Extracted {len(state['quantities'])} quantities, {len(state['materials'])} materials")
        
    except Exception as e:
        error_msg = f"Quantity extraction failed: {str(e)}"
        state["errors"].append(error_msg)
        print(f"❌ {error_msg}")
    
    return state


def analyze_competitors(state: TakeoffState) -> TakeoffState:
    """
    Stage 3: Scrape competitor websites using Firecrawl to identify best practices
    """
    print("🔍 Stage 3: Analyzing competitor takeoff solutions...")
    
    state["competitors_analyzed"] = []
    state["competitor_features"] = {}
    state["best_practices"] = []
    
    if not FIRECRAWL_API_KEY:
        print("⚠️  No Firecrawl API key - skipping competitor analysis")
        state["status"] = "competitors_analyzed"
        return state
    
    try:
        for competitor in COMPETITOR_TARGETS[:3]:  # Analyze top 3 to stay within limits
            try:
                print(f"  Scraping {competitor['name']}...")
                
                # Use Firecrawl to scrape competitor site
                response = requests.post(
                    "https://api.firecrawl.dev/v0/scrape",
                    headers={
                        "Authorization": f"Bearer {FIRECRAWL_API_KEY}",
                        "Content-Type": "application/json"
                    },
                    json={
                        "url": competitor["url"],
                        "pageOptions": {
                            "onlyMainContent": True
                        }
                    },
                    timeout=30
                )
                
                if response.status_code == 200:
                    data = response.json()
                    content = data.get("data", {}).get("markdown", "")
                    
                    # Analyze features using Claude
                    analysis = anthropic_client.messages.create(
                        model="claude-sonnet-4-20250514",
                        max_tokens=1000,
                        messages=[{
                            "role": "user",
                            "content": f"""Analyze this construction takeoff software and extract key features:

COMPETITOR: {competitor['name']}
CONTENT:
{content[:3000]}

Extract:
1. Key features (bullet points)
2. Automation capabilities
3. PDF processing capabilities
4. Reporting/export formats
5. Unique differentiators

Return JSON: {{"features": [...], "automation": [...], "differentiators": [...]}}"""
                        }]
                    )
                    
                    import re
                    json_match = re.search(r'\{.*\}', analysis.content[0].text, re.DOTALL)
                    if json_match:
                        features_data = json.loads(json_match.group())
                        state["competitor_features"][competitor["name"]] = features_data.get("features", [])
                        state["best_practices"].extend(features_data.get("differentiators", []))
                    
                    state["competitors_analyzed"].append(competitor["name"])
                    print(f"  ✅ {competitor['name']} analyzed")
                    
            except Exception as e:
                print(f"  ⚠️  Failed to analyze {competitor['name']}: {str(e)}")
                continue
        
        state["status"] = "competitors_analyzed"
        state["messages"].append({
            "timestamp": datetime.utcnow().isoformat(),
            "stage": "competitor_analysis",
            "message": f"Analyzed {len(state['competitors_analyzed'])} competitors"
        })
        
        print(f"✅ Competitor analysis complete: {len(state['competitors_analyzed'])} analyzed")
        
    except Exception as e:
        error_msg = f"Competitor analysis failed: {str(e)}"
        state["errors"].append(error_msg)
        print(f"❌ {error_msg}")
    
    return state


def calculate_costs(state: TakeoffState) -> TakeoffState:
    """
    Stage 4: Calculate cost estimates based on extracted quantities
    """
    print("💰 Stage 4: Calculating cost estimates...")
    
    try:
        cost_breakdown = {}
        labor_breakdown = {}
        total_cost = 0.0
        total_labor_hours = 0.0
        
        # Process each quantity
        for item in state["quantities"]:
            item_name = item.get("item", "").lower()
            quantity_str = str(item.get("quantity", "0"))
            category = item.get("category", "default")
            
            # Parse quantity
            try:
                quantity = float(quantity_str.replace("TBD", "0").replace(",", ""))
            except:
                quantity = 0.0
            
            # Find matching unit cost
            cost_data = None
            for key in UNIT_COSTS:
                if key in item_name or item_name in key:
                    cost_data = UNIT_COSTS[key]
                    break
            
            if not cost_data:
                cost_data = UNIT_COSTS["default"]
            
            # Calculate costs
            item_cost = quantity * cost_data["cost"]
            item_labor = quantity * cost_data["labor_hours"]
            
            cost_breakdown[item.get("item")] = {
                "quantity": quantity,
                "unit": cost_data["unit"],
                "unit_cost": cost_data["cost"],
                "total_cost": item_cost,
                "labor_hours": item_labor,
                "category": category
            }
            
            total_cost += item_cost
            total_labor_hours += item_labor
        
        # Add labor cost (assume $75/hour for Florida construction)
        labor_cost = total_labor_hours * 75.0
        
        state["cost_estimates"] = cost_breakdown
        state["labor_hours"] = {"total": total_labor_hours, "rate_per_hour": 75.0}
        state["total_estimated_cost"] = total_cost + labor_cost
        
        state["status"] = "costs_calculated"
        state["messages"].append({
            "timestamp": datetime.utcnow().isoformat(),
            "stage": "cost_calculation",
            "message": f"Total estimated cost: ${state['total_estimated_cost']:,.2f}"
        })
        
        print(f"✅ Costs calculated: ${state['total_estimated_cost']:,.2f} (Materials: ${total_cost:,.2f}, Labor: ${labor_cost:,.2f})")
        
    except Exception as e:
        error_msg = f"Cost calculation failed: {str(e)}"
        state["errors"].append(error_msg)
        print(f"❌ {error_msg}")
    
    return state


def generate_excel_report(state: TakeoffState) -> TakeoffState:
    """
    Stage 5: Generate comprehensive Excel takeoff report
    """
    print("📊 Stage 5: Generating Excel takeoff report...")
    
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # ====================================================================
        # SHEET 1: PROJECT SUMMARY
        # ====================================================================
        ws_summary = wb.create_sheet("Project Summary")
        
        # Header styling
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        
        # Title
        ws_summary["A1"] = "CONSTRUCTION TAKEOFF REPORT"
        ws_summary["A1"].font = Font(bold=True, size=16, color="1F4E78")
        ws_summary.merge_cells("A1:D1")
        
        # Project info
        ws_summary["A3"] = "Project Name:"
        ws_summary["B3"] = state.get("project_name", "Unnamed Project")
        ws_summary["A4"] = "Project ID:"
        ws_summary["B4"] = state.get("project_id", "N/A")
        ws_summary["A5"] = "Report Date:"
        ws_summary["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_summary["A6"] = "Drawing Type:"
        ws_summary["B6"] = state.get("drawing_type", "Unknown")
        ws_summary["A7"] = "Scale:"
        ws_summary["B7"] = state.get("scale_detected", "Not specified")
        
        # Cost summary
        ws_summary["A9"] = "COST SUMMARY"
        ws_summary["A9"].font = header_font
        ws_summary["A9"].fill = header_fill
        ws_summary.merge_cells("A9:D9")
        
        materials_cost = sum(item["total_cost"] for item in state["cost_estimates"].values())
        labor_cost = state["labor_hours"]["total"] * state["labor_hours"]["rate_per_hour"]
        
        ws_summary["A10"] = "Materials Cost:"
        ws_summary["B10"] = f"${materials_cost:,.2f}"
        ws_summary["A11"] = "Labor Cost:"
        ws_summary["B11"] = f"${labor_cost:,.2f}"
        ws_summary["A12"] = "Total Estimated Cost:"
        ws_summary["B12"] = f"${state['total_estimated_cost']:,.2f}"
        ws_summary["B12"].font = Font(bold=True, size=12)
        
        # ====================================================================
        # SHEET 2: DETAILED TAKEOFF
        # ====================================================================
        ws_takeoff = wb.create_sheet("Detailed Takeoff")
        
        # Headers
        headers = ["Item", "Quantity", "Unit", "Unit Cost", "Total Cost", "Labor Hours", "Category"]
        for col, header in enumerate(headers, start=1):
            cell = ws_takeoff.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        row = 2
        for item_name, item_data in state["cost_estimates"].items():
            ws_takeoff.cell(row=row, column=1).value = item_name
            ws_takeoff.cell(row=row, column=2).value = item_data["quantity"]
            ws_takeoff.cell(row=row, column=3).value = item_data["unit"]
            ws_takeoff.cell(row=row, column=4).value = item_data["unit_cost"]
            ws_takeoff.cell(row=row, column=5).value = item_data["total_cost"]
            ws_takeoff.cell(row=row, column=6).value = item_data["labor_hours"]
            ws_takeoff.cell(row=row, column=7).value = item_data["category"]
            row += 1
        
        # Format currency columns
        for row_idx in range(2, row):
            ws_takeoff.cell(row=row_idx, column=4).number_format = '$#,##0.00'
            ws_takeoff.cell(row=row_idx, column=5).number_format = '$#,##0.00'
        
        # Auto-fit columns
        for col in range(1, 8):
            ws_takeoff.column_dimensions[get_column_letter(col)].width = 15
        
        # ====================================================================
        # SHEET 3: MATERIALS LIST
        # ====================================================================
        ws_materials = wb.create_sheet("Materials List")
        
        headers = ["Material", "Quantity", "Unit", "Category", "Specification"]
        for col, header in enumerate(headers, start=1):
            cell = ws_materials.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        row = 2
        for material in state.get("materials", []):
            ws_materials.cell(row=row, column=1).value = material.get("material", "")
            ws_materials.cell(row=row, column=2).value = material.get("quantity", "")
            ws_materials.cell(row=row, column=3).value = material.get("unit", "")
            ws_materials.cell(row=row, column=4).value = material.get("category", "")
            ws_materials.cell(row=row, column=5).value = material.get("specification", "")
            row += 1
        
        # ====================================================================
        # SHEET 4: COMPETITOR INSIGHTS
        # ====================================================================
        if state.get("competitors_analyzed"):
            ws_comp = wb.create_sheet("Competitor Insights")
            
            ws_comp["A1"] = "COMPETITOR ANALYSIS"
            ws_comp["A1"].font = Font(bold=True, size=14, color="1F4E78")
            ws_comp.merge_cells("A1:C1")
            
            row = 3
            for comp_name, features in state.get("competitor_features", {}).items():
                ws_comp.cell(row=row, column=1).value = comp_name
                ws_comp.cell(row=row, column=1).font = Font(bold=True)
                row += 1
                
                for feature in features:
                    ws_comp.cell(row=row, column=2).value = f"• {feature}"
                    row += 1
                row += 1
            
            # Best practices
            ws_comp.cell(row=row, column=1).value = "BEST PRACTICES IDENTIFIED:"
            ws_comp.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            for practice in state.get("best_practices", []):
                ws_comp.cell(row=row, column=2).value = f"✓ {practice}"
                row += 1
        
        # Save workbook
        output_dir = Path("/tmp/spd_takeoff_reports")
        output_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"takeoff_{state['project_id']}_{timestamp}.xlsx"
        excel_path = output_dir / excel_filename
        
        wb.save(str(excel_path))
        
        state["excel_path"] = str(excel_path)
        state["status"] = "report_generated"
        
        # Generate summary
        state["report_summary"] = f"""
Construction Takeoff Report Generated
=====================================

Project: {state['project_name']}
Items: {len(state['quantities'])} quantities
Materials: {len(state['materials'])} materials
Total Cost: ${state['total_estimated_cost']:,.2f}

Excel Report: {excel_path}
"""
        
        state["messages"].append({
            "timestamp": datetime.utcnow().isoformat(),
            "stage": "report_generation",
            "message": f"Excel report generated: {excel_filename}"
        })
        
        print(f"✅ Excel report generated: {excel_path}")
        
    except Exception as e:
        error_msg = f"Report generation failed: {str(e)}"
        state["errors"].append(error_msg)
        state["status"] = "failed"
        print(f"❌ {error_msg}")
    
    return state


def log_to_supabase(state: TakeoffState) -> TakeoffState:
    """
    Stage 6: Log results to Supabase insights table
    """
    print("💾 Stage 6: Logging to Supabase...")
    
    if not SUPABASE_KEY:
        print("⚠️  No Supabase key - skipping database logging")
        state["status"] = "completed"
        return state
    
    try:
        insight_data = {
            "category": "spd_construction_takeoff",
            "insight_type": "takeoff_report",
            "insight_data": {
                "project_id": state["project_id"],
                "project_name": state["project_name"],
                "drawing_type": state["drawing_type"],
                "total_quantities": len(state["quantities"]),
                "total_materials": len(state["materials"]),
                "total_cost": state["total_estimated_cost"],
                "labor_hours": state["labor_hours"]["total"],
                "competitors_analyzed": state["competitors_analyzed"],
                "excel_report": state["excel_path"],
                "status": state["status"],
                "timestamp": datetime.utcnow().isoformat()
            }
        }
        
        response = requests.post(
            f"{SUPABASE_URL}/rest/v1/insights",
            headers={
                "apikey": SUPABASE_KEY,
                "Authorization": f"Bearer {SUPABASE_KEY}",
                "Content-Type": "application/json",
                "Prefer": "return=minimal"
            },
            json=insight_data,
            timeout=10
        )
        
        if response.status_code in [200, 201]:
            print("✅ Logged to Supabase")
        else:
            print(f"⚠️  Supabase logging failed: {response.status_code}")
        
    except Exception as e:
        print(f"⚠️  Supabase logging error: {str(e)}")
    
    state["status"] = "completed"
    return state

# ============================================================================
# LANGGRAPH WORKFLOW CONSTRUCTION
# ============================================================================

def create_takeoff_workflow() -> StateGraph:
    """
    Create LangGraph workflow for construction takeoff
    """
    workflow = StateGraph(TakeoffState)
    
    # Add nodes
    workflow.add_node("process_pdf", process_pdf)
    workflow.add_node("extract_quantities", extract_quantities)
    workflow.add_node("analyze_competitors", analyze_competitors)
    workflow.add_node("calculate_costs", calculate_costs)
    workflow.add_node("generate_report", generate_excel_report)
    workflow.add_node("log_results", log_to_supabase)
    
    # Define edges
    workflow.add_edge("process_pdf", "extract_quantities")
    workflow.add_edge("extract_quantities", "analyze_competitors")
    workflow.add_edge("analyze_competitors", "calculate_costs")
    workflow.add_edge("calculate_costs", "generate_report")
    workflow.add_edge("generate_report", "log_results")
    workflow.add_edge("log_results", END)
    
    # Set entry point
    workflow.set_entry_point("process_pdf")
    
    return workflow.compile()

# ============================================================================
# MAIN EXECUTION FUNCTION
# ============================================================================

def run_construction_takeoff(pdf_path: str, project_name: str, project_id: str = None) -> TakeoffState:
    """
    Execute construction takeoff workflow
    
    Args:
        pdf_path: Path to construction plan PDF
        project_name: Name of the construction project
        project_id: Optional project identifier
    
    Returns:
        TakeoffState with results
    """
    # Initialize state
    if not project_id:
        project_id = f"TKO-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    
    initial_state = TakeoffState(
        pdf_path=pdf_path,
        pdf_base64="",
        project_name=project_name,
        project_id=project_id,
        pdf_extracted_text="",
        pdf_pages_analyzed=0,
        drawing_type="",
        scale_detected="",
        quantities=[],
        materials=[],
        measurements={},
        competitors_analyzed=[],
        competitor_features={},
        best_practices=[],
        cost_estimates={},
        labor_hours={},
        total_estimated_cost=0.0,
        excel_path="",
        report_summary="",
        messages=[],
        errors=[],
        status="initiated"
    )
    
    # Create and run workflow
    print(f"\n🚀 Starting Construction Takeoff: {project_name} ({project_id})")
    print("=" * 80)
    
    workflow = create_takeoff_workflow()
    final_state = workflow.invoke(initial_state)
    
    print("\n" + "=" * 80)
    print(f"✅ Takeoff Complete: {final_state['status']}")
    print(final_state['report_summary'])
    
    return final_state


# ============================================================================
# CLI INTERFACE
# ============================================================================

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 3:
        print("Usage: python construction_takeoff_agent.py <pdf_path> <project_name> [project_id]")
        sys.exit(1)
    
    pdf_path = sys.argv[1]
    project_name = sys.argv[2]
    project_id = sys.argv[3] if len(sys.argv) > 3 else None
    
    result = run_construction_takeoff(pdf_path, project_name, project_id)
    
    print(f"\nExcel Report: {result['excel_path']}")
    
    if result['errors']:
        print("\n⚠️  Errors encountered:")
        for error in result['errors']:
            print(f"  - {error}")
